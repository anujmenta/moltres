import pandas as pd
import re
from ratecard import *
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Border, Side
from openpyxl.styles import Alignment
from math import ceil
import argparse
import json

parser = argparse.ArgumentParser(description='Optional app description')
parser.add_argument('--input', type=str,
                    help='Path to csv bill')
parser.add_argument('--output', type=str,
                    help='Path to output report')
parser.add_argument('--subcats', type=str, help='Sub categories boolean values')

args = parser.parse_args()

machines = pd.read_csv('TCO reference - C-Machines.csv')
rates = pd.read_csv('TCO reference - O-Rates.csv')

# csv_to_read = input('path to csv file: ')
# df = pd.read_csv(csv_to_read)



if args.subcats:
  subcats_args = eval(args.subcats)
  category_toggle = {
    'suds_bool' : subcats_args['suds_bool'],
    'box_compute': subcats_args['box_compute'],
    'heavy_compute' : subcats_args['heavy_compute'],
    'spot_compute' : subcats_args['spot_compute'],
    'persistentdisk' : subcats_args['persistentdisk'],
    'cloudstorage':subcats_args['cloudstorage'],
    'loadbalancer' : subcats_args['loadbalancer'],
    'cloudnat' : subcats_args['cloudnat'],
    'idleaddress' : subcats_args['idleaddress'],
    'cloudsql' : subcats_args['cloudsql'],
    'egress' : subcats_args['egress'],
    'support' : subcats_args['support'],
  }
else:
  category_toggle = {
    'suds_bool' : True,
    'box_compute': True,
    'heavy_compute' : True,
    'spot_compute' : True,
    'persistentdisk' : True,
    'cloudstorage':True,
    'loadbalancer' : True,
    'cloudnat' : True,
    'idleaddress' : True,
    'cloudsql' : True,
    'egress' : True,
    'support' : True,
  }

df = pd.read_csv(args.input)
recordtype_options = {
  'lineItem/LineItemType' : 'Usage',
  'RecordType' : ['LineItem', 'PayerLineItem', 'LinkedLineItem'],
}
try:
  recordtype_col = list(set(recordtype_options).intersection(set(df.columns)))[0]
  df = df[(df[recordtype_col].isin(recordtype_options[recordtype_col]))|(df[recordtype_col]=='RIFee')]
except:
  print('Looks like Recordtype is not available in args.input')
  pass
#function to detect column names
def detect_column_names(dataframe):
  usage_options = ['lineItem/UsageType', 'UsageType']
  cost_options = ['lineItem/UnblendedCost', 'TotalCost', "Cost"]
  rate_options = ['BlendedRate', 'UnblendedRate', 'lineItem/UnblendedRate', "Rate"]
  description_options = ['lineItem/LineItemDescription', 'ItemDescription']
  quantity_options = ['lineItem/UsageAmount', 'UsageQuantity']
  productname_options = ['product/ProductName', 'ProductCode', "ProductName"]
  type_options = ['ItemType']

  columnnames = {
  'usage': list(set(usage_options).intersection(set(dataframe.columns)))[0],
  'cost': list(set(cost_options).intersection(set(dataframe.columns)))[0],
  'rate': list(set(rate_options).intersection(set(dataframe.columns)))[0],
  'description': list(set(description_options).intersection(set(dataframe.columns)))[0],
  'quantity': list(set(quantity_options).intersection(set(dataframe.columns)))[0],
  'productname': list(set(productname_options).intersection(set(dataframe.columns)))[0],
  # 'type': set(type_options).intersection(set(dataframe.columns)))[0],
  }
  return columnnames


#Main component
def sud(nunits, rate):
  n30, remaining = int(nunits)//744, int(nunits)%744
  n15, remaining = remaining//360, remaining%360
  h30, h15, h0 = n30*744, n15*360, remaining
  final = (0.7*h30/nunits + 0.85*h15/nunits + remaining/nunits)*rate
  return final#Returns normalized rate after sud

ratedict = {
    'box': 'List-price',
    'heavy': 'Commitment-1yr',
    'spot': 'Preemptible-use'
}

ssd_rate = {
    'box' : 0.08,
    'heavy' : 0.051,
    'spot' : 0.048
}

gpu_rates = {
    'NVIDIA Tesla K80':{
        'box' : 0.45,
        'heavy' : 0.283, 
        'spot' : 0.135
    },
    'NVIDIA Tesla V100':{
        'box' : 2.48,
        'heavy' : 1.562,
        'spot' : 0.74,
    },
    # 'NVIDIA GRID K520':{
    #     'box' : 2.48,
    #     'heavy' : 1.562,
    #     'spot' : 0.74,
    # },
    # 'NVIDIA Tesla M60':{
    #     'box' : 2.48,
    #     'heavy' : 1.562,
    #     'spot' : 0.74,
    # },
    # 'NVIDIA T4 Tensor Core':{
    #     'box' : 2.48,
    #     'heavy' : 1.562,
    #     'spot' : 0.74,
    # }
}


testmode = False

sud_dict = {
    'GPU' : True,
    'All' : True,
}

def get_usagetype(row):

  val = row[columnnames['usage']]

  if 'BoxUsage' in val:
    usagetype = 'box'
  elif 'HeavyUsage' in val:
    usagetype = 'heavy'
  elif 'SpotUsage' in val:
    usagetype = 'spot'
  # print(val)

  if row[columnnames['productname']]=='Amazon Elastic MapReduce':
    usagetype = 'spot'

  return usagetype

def get_region(usagetypeval):
  """
  Takes the 'UsageType' value of a rowitem and returns the GCP and AWS regions """
  try:
    region_aws, rest = usagetypeval.split('-')
  except:
    region_aws, rest = 'USE1', usagetypeval
    # return pd.Series([0]*8)
  region_gcp = region_dict[region_aws]

  return (region_gcp, region_aws, rest)

def parsecompute(row):
  val = row[columnnames['usage']]
  compute_status = 'Manual'
  usagetype = get_usagetype(row)
  # print(val)
  try:
    region_aws, rest = val.split('-')
  except:
    region_aws, rest = 'USE1', val
    # return pd.Series([0]*8)
  region_gcp = region_dict[region_aws]
  try:
    usage, aws_machine = rest.split(':')
  except:
    usage, aws_machine = rest, ''
    return pd.Series([0]*7+[compute_status])
  machine = aws_machine.replace('.elasticsearch', '')
  if machine not in list(machines['API Name']):
    return pd.Series([0]*7+[compute_status])
  # print(machine)
  gcpmachine = machines[machines['API Name']==machine].iloc[0]
  gcpmachine_name = [gcpmachine['Family'].upper(), gcpmachine['Type'].capitalize(), gcpmachine['Memory'], gcpmachine['vCPUs']]
  if gcpmachine_name[1]=='Standard' or gcpmachine_name[1]=='Highmem' or gcpmachine_name[1]=='Highcpu':
    gcpmachine_name[1] = 'Predefined'
  #Needs restructuring to accommodate for all regions and pull rate from the new JSON
  if 'windows' in row[columnnames['description']].lower():
    os_type = 'win'
  elif 'suse' in row[columnnames['description']].lower():
    os_type = 'suse'
  else:
    os_type = ''
  if gcpmachine_name[1] not in ['Small', 'Micro', 'Medium']:
    machine_id = '{}-{}'.format(gcpmachine_name[0],gcpmachine_name[1]).lower()
    rates_cpu = compute_ratecard[usagetype]['{}-vcpus'.format(machine_id)][region_gcp]
    rates_mem = compute_ratecard[usagetype]['{}-memory'.format(machine_id)][region_gcp]
    gcp_rate = gcpmachine_name[2]*float(rates_mem)+gcpmachine_name[3]*float(rates_cpu)
    if machine.startswith('p'):
      if gcpmachine['GPU model']=='NVIDIA Tesla K80':
        gcp_rate+=int(gcpmachine['GPUs'])*gpu_rates['NVIDIA Tesla K80'][usagetype]
      elif gcpmachine['GPU model']=='NVIDIA Tesla V100':
        gcp_rate+=int(gcpmachine['GPUs'])*gpu_rates['NVIDIA Tesla V100'][usagetype]
    if os_type:
      os_rate = compute_ratecard['os'][os_type]['high']
    else:
      os_rate = 0
    gcp_rate+=os_rate*gcpmachine_name[3]
  else:
    machine_id = '{}-{}'.format(gcpmachine_name[0],gcpmachine_name[1]).lower()
    gcp_rate = compute_ratecard[usagetype][machine_id][region_gcp]
    if os_type:
      os_rate = compute_ratecard['os'][os_type]['low']
    else:
      os_rate = 0
    gcp_rate+=os_rate*gcpmachine_name[3]
  original_rate = gcp_rate
  if row[columnnames['rate']]:
    nunits = row[columnnames['cost']]/row[columnnames['rate']]
    if usagetype=='box' and category_toggle['suds_bool']:
      gcp_rate = sud(nunits, gcp_rate)
  else:
    nunits = 0
  ssd_cost = 0
  if gcpmachine['Instance Storage']!=0 and gcpmachine['SSD/HDD?']=='SSD':
    ssd_cost +=int(gcpmachine['Instance Storage'])*ssd_rate[usagetype]*(nunits//744)
  sud_savings = nunits*(original_rate-gcp_rate)
  # print(ssd_cost, int(gcpmachine['Instance Storage']), ssd_rate[usagetype], nunits, nunits//744)
  # print(nunits, original_rate)
  compute_status = 'Done'
  returner = [nunits, '{}, {}cpus, {} mem'.format(machine_id, gcpmachine_name[3], gcpmachine_name[2])]+[gcp_rate, nunits*gcp_rate+ssd_cost, sud_savings, ssd_cost, sud_savings/(nunits*original_rate)*100, compute_status]
  return pd.Series(returner)

columnnames = detect_column_names(df)

# for key in columnnames:
# 	columnnames[key] = input('Enter column name for {}: '.format(key))

df = df[df[columnnames['cost']]>0]
df = df[df[columnnames['usage']]!='']
# df = df[~df[columnnames['description']].str.contains('Total')]

df[columnnames['rate']] = df[columnnames['cost']]/df[columnnames['quantity']]
grouped = df.groupby([columnnames['usage'], columnnames['description'], columnnames['productname']]).agg({columnnames['cost']:'sum', columnnames['rate']:'mean', columnnames['quantity']:'sum'}).reset_index()

aws_actual_cost = sum(grouped[columnnames['cost']])
print('AWS total cost: ', aws_actual_cost)

# # print(boxusage)
# if len(boxusage):
#   print('Box AWS: ', sum(boxusage[columnnames['cost']]))
  
#   # print('Boxusage', sum(boxusage[columnnames['cost']]), sum(boxusage['gcp_cost']))
# if len(heavyusage):
#   print('heavy AWS: ', sum(heavyusage[columnnames['cost']]), len(heavyusage))
  
#   # print('Heavyusage', sum(heavyusage[columnnames['cost']]), sum(heavyusage['gcp_cost']))
# if len(spotusage):
#   print('spot AWS: ', sum(spotusage[columnnames['cost']]))
  
  # print('Spotusage', sum(spotusage[columnnames['cost']]), sum(spotusage['gcp_cost']))

######################################################################################################################################################

def parsepd(row):
  compute_status = 'Manual'
  value = row[columnnames['usage']]
  # print(value)
  try:
    if not value.startswith('EBS'):
      region, cat = value.split('-EBS:')
    else:
      region, cat = 'USE1', value.replace('EBS:', '')
    if cat=='SnapshotUsage':
      compute_status = 'Done'
      return pd.Series([persistentdisk_ratecard['snapshot'][region_dict[region]], row[columnnames['quantity']]*persistentdisk_ratecard['snapshot'][region_dict[region]], compute_status])
    elif cat=='VolumeUsage.gp2':
      compute_status = 'Done'
      return pd.Series([persistentdisk_ratecard['gp2'][region_dict[region]], row[columnnames['quantity']]*persistentdisk_ratecard['gp2'][region_dict[region]], compute_status])
    elif 'Magnetic provisioned storage' in row[columnnames['description']]:
      compute_status = 'Done'
      return pd.Series([persistentdisk_ratecard['magnetic'][region_dict[region]], row[columnnames['quantity']]*persistentdisk_ratecard['magnetic'][region_dict[region]], compute_status])
  except:
    return pd.Series([0, 0, compute_status])


######################################################################################################################################################

def nat_gateway_cost(dataframe):
  nat_gateway_bytes = dataframe[dataframe[columnnames['usage']].str.contains('NatGateway-Bytes')].copy()
  nat_gateway_hours = dataframe[dataframe[columnnames['usage']].str.contains('NatGateway-Hours')].copy()
  nat_gateway_hours['ninstances'] = nat_gateway_hours[columnnames['quantity']]/720
  ninstances = sum(nat_gateway_hours['ninstances'])
  ngb = sum(nat_gateway_bytes[columnnames['quantity']])
  if ninstances<=32:
    cost = ninstances*nat_gateway_ratecard['us-vm-low']*720+ngb*nat_gateway_ratecard['us-bytes']
  else:
    cost = nat_gateway_ratecard['us-vm-high']*720+ngb*nat_gateway_ratecard['us-bytes']
  return [cost, sum(nat_gateway_bytes[columnnames['cost']])+sum(nat_gateway_hours[columnnames['cost']])]


#feed this input into the viz tool
######################################################################################################################################################

def idle_addresses_cost(row):
  compute_status = 'Manual'
  ninstances = int(row[columnnames['quantity']]/720)
  compute_status = 'Done'
  return pd.Series([ninstances*idle_addresses_ratecard['us']*720, row[columnnames['cost']], compute_status])

######################################################################################################################################################

def loadbalancer_cost(row):
  rowsplit = row[columnnames['usage']].split('-')
  region_gcp = region_dict[rowsplit[0]]
  usecase = '-'.join(rowsplit[1:])
  return pd.Series([region_gcp, usecase])

def loadgrouped_cost(row):
  compute_status = 'Manual'
  region = row['region']
  usecase = row['usecase']
  quantity = row[columnnames['quantity']]
  if usecase=='LoadBalancerUsage':
    nrules = ceil(quantity/720)
    cost = loadbalancer_ratecard['forwarding_rules'][region]*720
    if nrules>5:
      cost+=(nrules-5)*loadbalancer_ratecard['forwarding_rules_extra'][region]*720
    compute_status = 'Done'
    return pd.Series([nrules, cost, compute_status])
  else:
    cost = quantity*loadbalancer_ratecard['ingress'][region]
    compute_status = 'Done'
    return pd.Series([quantity, cost, compute_status])

######################################################################################################################################################

def cloudstorage_cost(row):
  usageval = row[columnnames['usage']]
  compute_status = 'Manual'
  desc = row[columnnames['description']]
  quantity = row[columnnames['quantity']]
  if 'Monitoring-Automation' in usageval or 'StorageAnalytics' in usageval or 'TagStorage' in usageval:
    return pd.Series(['','', quantity, 0,0, compute_status])

  if usageval.endswith('Tier1') or usageval.endswith('Tier4'):
    gcp_sku = 'class-a'
  elif usageval.endswith('Tier2') or usageval.endswith('Tier3'):
    gcp_sku = 'class-b'
  elif 'Retrieval' in usageval:
    gcp_sku = 'retrieval'
  elif 'TimedStorage' in usageval:
    gcp_sku = 'storage'
  elif 'EarlyDelete' in usageval:
    gcp_sku = 'storage'
  else:
    # print(row)
    return pd.Series(['','', quantity, 0,row[columnnames['cost']], compute_status])
  if 'Glacier Deep Archive' in desc:
    gcp_class = 'archive'
  elif 'Glacier' in desc:
    gcp_class = 'coldline'
  elif 'Infrequent Access' in desc:
    gcp_class = 'nearline'
  else:
    gcp_class = 'standard'
  
  region_aws = usageval.split('-')[0]
  if region_aws in region_dict:
    region_gcp = region_dict[region_aws]
  else:
    region_gcp = region_dict['USE1']
  gcp_rate = cloudstorage_ratecard['{}-{}'.format(gcp_sku, gcp_class)][region_gcp]
  if gcp_sku in ['class-a', 'class-b']:
    gcp_rate/=10000
  gcp_cost = gcp_rate*quantity
  compute_status = 'Done'
  return pd.Series([gcp_class, gcp_sku, quantity, gcp_rate, gcp_cost, compute_status])


######################################################################################################################################################

def parsecloudsql(row):
  compute_status = 'Manual'
  usagevalue = row[columnnames['usage']]
  quantity = row[columnnames['quantity']]
  desc = row[columnnames['description']] 
  ha_multiplier = 1
  if 'Multi-AZ' in usagevalue:
    ha_multiplier = 2
  if 'MySQL' in desc:
    engine = 'MySQL'
  elif 'PostgreSQL' in desc:
    engine = 'PostgreSQL'
  elif 'SQL Server' in desc:
    engine = 'SQL Server'
  else:
    engine = 'MySQL'
  # gcp_region, aws_region, rest = get_region(usagevalue)
  
  try:
    rest, machine = usagevalue.split(':')
  except:
    rest, machine = usagevalue, ''
    return pd.Series([0]*2+[compute_status])
  # print(rest, machine)
  try:
    if '-' in rest:
      region_aws = rest.split('-')[0]
      usage = rest[len(region_aws)+1:]
    else:
      region_aws, usage = 'USE1', rest
  except:
    region_aws, usage = 'USE1', rest
  # try:
  #   region_aws, usage = rest.split('-')
  # except:
  #   region_aws, usage = 'USE1', rest
    # return pd.Series([0]*8)
  gcp_region = region_dict[region_aws]
  # print(usage)
  # print(region_aws, gcp_region, usage, machine)
  if usage=='InstanceUsage' or usage=='Multi-AZUsage':
    if machine in cloudsql_machineref:
      gcp_machine = cloudsql_machineref[machine]
      gcp_rate = cloudsql_instances_ratecard['CP-'+gcp_machine.upper()][gcp_region]*ha_multiplier
      compute_status = 'Done'
      return pd.Series([gcp_rate, gcp_rate*quantity, compute_status])
    else:
      # print('Machine not found, ', machine)
      return pd.Series([0]*2+[compute_status])
  elif usage=='RDS' and 'GP2-Storage' in machine:
    gcp_rate = cloudsql_other_ratecard['gp2'][gcp_region]
    compute_status = 'Done'
    return pd.Series([gcp_rate, gcp_rate*quantity, compute_status])
  else:
    # print('Usage not found, ',usage)
    return pd.Series([0]*2+[compute_status])

######################################################################################################################################################


def gbsplitter(val):
  if val>143360:
    return [10240, 10240*13, val-143360]
  elif val>10240:
    return [10240, val-10240, 0]
  else:
    return [val, 0, 0]

def egress_cost(row):
  # print(row)
  compute_status = 'Manual'
  usageval = row[columnnames['usage']]
  quantity = row[columnnames['quantity']]
  try:
    if 'DataTransfer-Regional-Bytes' in usageval:
      aws_region = usageval.replace('DataTransfer-Regional-Bytes', '')
      if aws_region:
        aws_region = usageval.split('-')[0]
      else:
        aws_region = 'USW2'
      gcp_region = region_dict[aws_region]
      gcp_rate = network_egress_ratecard['vm-vm-internal'][gcp_region.split('-')[0]][gcp_region]
      gcp_cost = quantity*gcp_rate
    elif 'DataTransfer' in usageval:
      aws_region = usageval.replace('DataTransfer-Out-Bytes', '')
      if aws_region:
        aws_region = usageval.split('-')[0]
      else:
        aws_region = 'USW2'
      gcp_region = region_dict[aws_region]
      gcp_rate = network_egress_ratecard['to-worldwide'][gcp_region.split('-')[0]]
      gcp_rate = [gcp_rate["10240"], gcp_rate["143360"], gcp_rate["143361"]]
      gcp_cost = sum(x*y for x,y in zip(gcp_rate, gbsplitter(quantity)))
      gcp_rate = ' '.join(map(str, gcp_rate)) # To let it convert peacefully to excel
    elif 'AWS-Out-Bytes' in usageval:
      # print(row)
      aws_regions = usageval.replace('-AWS-Out-Bytes', '').split('-')
      if len(aws_regions)==2:
        gcp_region_from = region_dict[aws_regions[0]]
        gcp_region_to = region_dict[aws_regions[1]]
        # print(usageval)
        gcp_rate = network_egress_ratecard['vm-vm-external']['{}-{}'.format(gcp_region_from.split('-')[0], gcp_region_to.split('-')[0])][gcp_region_from]
        gcp_cost = gcp_rate*quantity
      else:
        gcp_rate, gcp_cost = 0,0
    compute_status = 'Done'
    return pd.Series([gcp_rate, gcp_cost, compute_status])
  except:
    return pd.Series([0,0, compute_status])


# cost_matrix['aws']['egress'] = round(sum(egress_df[columnnames['cost']]), 2)
# cost_matrix['aws']['egress'] = round(sum(egress_df['gcp_cost']), 2)
# print(cost_matrix['aws']['egress'], cost_matrix['aws']['egress'])

######################################################################################################################################################


cost_matrix = {
  'aws':{
    'box_compute' : 0,
    'heavy_compute' : 0,
    'spot_compute' : 0,
    'persistentdisk' : 0,
    'cloudstorage': 0,
    'loadbalancer' : 0,
    'cloudnat' : 0,
    'idleaddress' : 0,
    'egress': 0,
    'cloudsql': 0, 
    'support': 0,
  },
  'gcp':{
    'box_compute': 0,
    'heavy_compute' : 0,
    'spot_compute' : 0,
    'persistentdisk' : 0,
    'cloudstorage':0,
    'loadbalancer' : 0,
    'cloudnat' : 0,
    'idleaddress' : 0,
    'egress': 0,
    'cloudsql': 0, 
    'support': 0,

  },
}

def tcocomputer():
  box_filter = grouped[columnnames['usage']].str.contains('BoxUsage')
  boxusage = grouped[box_filter].sort_values(columnnames['usage'])
  if len(boxusage) and category_toggle['box_compute']:
    boxusage[['nunits', 'gcp_type', 'gcp_rate', 'gcp_cost', 'sud_savings', 'ssd_cost', 'sud_savings_percentage', 'Status']] = boxusage.apply(parsecompute, axis=1)
    cost_matrix['aws']['box_compute'] = round(sum(boxusage[columnnames['cost']]), 2)
    cost_matrix['gcp']['box_compute'] = round(sum(boxusage['gcp_cost']), 2)
    grouped = grouped[~box_filter]

  heavy_filter = (grouped[columnnames['usage']].str.contains('HeavyUsage'))&(~grouped[columnnames['usage']].str.contains('HeavyUsage:db')&((grouped[columnnames['productname']]=='Amazon Elastic Compute Cloud')|(grouped[columnnames['productname']]=='AmazonEC2')))
  heavyusage = grouped[heavy_filter].sort_values(columnnames['usage'])
  if len(heavyusage) and category_toggle['heavy_compute']:
    heavyusage[['nunits', 'gcp_type', 'gcp_rate', 'gcp_cost', 'sud_savings', 'ssd_cost', 'sud_savings_percentage', 'Status']] = heavyusage.apply(parsecompute, axis=1)
    cost_matrix['aws']['heavy_compute'] = round(sum(heavyusage[columnnames['cost']]), 2)
    cost_matrix['gcp']['heavy_compute'] = round(sum(heavyusage['gcp_cost']), 2)
    grouped = grouped[~heavy_filter]

  spot_filter = grouped[columnnames['usage']].str.contains('SpotUsage')
  spotusage = grouped[spot_filter].sort_values(columnnames['usage'])
  if len(spotusage) and category_toggle['spot_compute']:
    spotusage[['nunits', 'gcp_type', 'gcp_rate', 'gcp_cost', 'sud_savings', 'ssd_cost', 'sud_savings_percentage', 'Status']] = spotusage.apply(parsecompute, axis=1)
    cost_matrix['aws']['spot_compute'] = round(sum(spotusage[columnnames['cost']]), 2)
    cost_matrix['gcp']['spot_compute'] = round(sum(spotusage['gcp_cost']), 2)
    grouped = grouped[~spot_filter]

  pd_filter = (grouped[columnnames['usage']].str.contains('VolumeUsage.gp2')) | (grouped[columnnames['usage']].str.contains('SnapshotUsage')) | ((grouped[columnnames['usage']].str.contains('VolumeUsage'))&(grouped[columnnames['description']].str.contains('Magnetic provisioned storage')))
  persistentdisk = grouped[pd_filter].sort_values(columnnames['usage'])
  if len(persistentdisk) and category_toggle['persistentdisk']:
    persistentdisk[['GCP_rate', 'GCP_cost', 'Status']] = persistentdisk.apply(parsepd, axis=1)
    cost_matrix['aws']['persistentdisk'] = round(sum(persistentdisk[columnnames['cost']]), 2)
    cost_matrix['gcp']['persistentdisk'] = round(sum(persistentdisk['GCP_cost']), 2)
    grouped = grouped[~pd_filter]
    print("PD AWS: ", sum(persistentdisk[columnnames['cost']]), " PD GCP: ", sum(persistentdisk['GCP_cost']))

  cloudstorage_filter = (grouped[columnnames['productname']]=='AmazonS3GlacierDeepArchive') | (grouped[columnnames['productname']]=='AmazonS3')
  cloudstorage = grouped[cloudstorage_filter]
  if len(cloudstorage) and category_toggle['cloudstorage']:
    cloudstorage[['GCP Class', 'GCP SKU', 'Quantity', 'GCP Rate', 'GCP Cost', 'Status']] = cloudstorage.apply(cloudstorage_cost, axis=1)
    cost_matrix['aws']['cloudstorage'] = round(sum(cloudstorage[columnnames['cost']]), 2)
    cost_matrix['gcp']['cloudstorage'] = round(sum(cloudstorage['GCP Cost']), 2)
    grouped = grouped[~cloudstorage_filter]

  loadbalancer_filter = ((grouped[columnnames['usage']].str.contains('DataProcessing-Bytes'))|(grouped[columnnames['usage']].str.contains('LCUUsage'))|(grouped[columnnames['usage']].str.contains('LoadBalancerUsage')))&(grouped[columnnames['productname']]=='Amazon Elastic Compute Cloud')
  loadbalancer = grouped[loadbalancer_filter]
  if len(loadbalancer) and category_toggle['loadbalancer']:
    loadbalancer[['region', 'usecase', 'Status']] = loadbalancer.apply(loadbalancer_cost, axis=1)
    load_grouped = loadbalancer.groupby(['region', 'usecase']).sum().reset_index()[['region', 'usecase', columnnames['quantity']]]
    load_grouped[['rules/gb', 'gcp_cost', 'Status']] = load_grouped.apply(loadgrouped_cost, axis=1)
    cost_matrix['aws']['loadbalancer'] = round(sum(loadbalancer[columnnames['cost']]), 2)
    cost_matrix['gcp']['loadbalancer'] = round(sum(load_grouped['gcp_cost']), 2)
    grouped = grouped[~loadbalancer_filter]

  nat_filter = grouped[columnnames['usage']].str.contains('NatGateway-Bytes') | grouped[columnnames['usage']].str.contains('NatGateway-Hours')
  nat_df = grouped[nat_filter]
  if len(nat_df) and category_toggle['cloudnat']:
    nat_gcp, nat_aws = nat_gateway_cost(nat_df)
    cost_matrix['aws']['cloudnat'] = round(nat_aws, 2)
    cost_matrix['gcp']['cloudnat'] = round(nat_gcp, 2)
    grouped = grouped[~nat_filter]

  idle_filter = grouped[columnnames['usage']].str.contains('ElasticIP:IdleAddress')
  idleaddress = grouped[idle_filter]
  if len(idleaddress) and category_toggle['idleaddress']:
    idleaddress[['gcp_cost', 'aws_cost', 'Status']] = idleaddress.apply(idle_addresses_cost, axis=1)
    cost_matrix['aws']['idleaddress'] = round(sum(idleaddress['aws_cost']), 2)
    cost_matrix['gcp']['idleaddress'] = round(sum(idleaddress['gcp_cost']), 2)
    grouped = grouped[~idle_filter]

  cloudsql_filter = (grouped[columnnames['productname']]=='Amazon Relational Database Service') | (grouped[columnnames['productname']]=='AmazonRDS')
  cloudsql = grouped[cloudsql_filter]
  if len(cloudsql) and category_toggle['cloudsql']:
    cloudsql[['gcp_rate', 'gcp_cost', 'Status']] = cloudsql.apply(parsecloudsql, axis=1)
    cost_matrix['aws']['cloudsql'] = round(sum(cloudsql[columnnames['cost']]), 2)
    cost_matrix['gcp']['cloudsql'] = round(sum(cloudsql['gcp_cost']), 2)
    grouped = grouped[~cloudsql_filter]

  egress_filter = (grouped[columnnames['usage']].str.contains('DataTransfer'))|(grouped[columnnames['usage']].str.contains('Out-Bytes'))
  egress_df = grouped[egress_filter]
  egress_df = egress_df.groupby(columnnames['usage']).sum().reset_index()
  # egress_df = egress_df[egress_df[columnnames['cost']]!=0]
  if len(egress_df) and category_toggle['egress']:
    egress_df[['gcp_rate', 'gcp_cost', 'Status']] = egress_df.apply(egress_cost, axis=1)
    cost_matrix['aws']['egress'] = round(sum(egress_df[columnnames['cost']]), 2)
    cost_matrix['gcp']['egress'] = round(sum(egress_df['gcp_cost']), 2)
    grouped = grouped[~egress_filter]

  support_filter = grouped[columnnames['usage']]=='Dollar'
  support = grouped[support_filter]
  if len(support) and category_toggle['support']:
    cost_matrix['aws']['support'] = round(sum(support[columnnames['cost']]), 2)
    cost_matrix['gcp']['support'] = 0
    grouped = grouped[~support_filter]

  grouped = grouped.sort_values(columnnames['cost'], ascending=False)
# print(cost_matrix)
######################################################################################################################################################
# support_filter = grouped[columnnames['usage']]=='Dollar'
# support = grouped[support_filter]
# grouped = grouped[~support_filter]


######################################################################################################################################################

def reporting():
  wb = Workbook()

  report_filename = args.output

  ws_summary = wb.active
  ws_summary.title = 'Summary'

  ws_summary.append([''])
  ws_summary.append(['', 'Summarized View - Total Cost of Ownership on GCP'])
  ws_summary.append(['', 'Type', 'GCP', 'Estimates', 'AWS', 'Actuals', '% Savings'])
  center = ws_summary['B2']
  center.alignment = Alignment(horizontal='center')
  ws_summary.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
  ws_summary.append(['', 'Compute', 'GCE - Preemptible VMs', '${}'.format(cost_matrix['gcp']['spot_compute']), 'EC2 - Spot Usage', '${}'.format(cost_matrix['aws']['spot_compute'])])
  ws_spot = wb.create_sheet(title='SpotUsage')

  for column_cells in ws_summary.columns:
    ws_summary.column_dimensions[column_cells[0].column_letter].width = 15

  ws_summary.column_dimensions[ws_summary.cell(row=ws_summary._current_row, column=3).column_letter].width = 20
  ws_summary.column_dimensions[ws_summary.cell(row=ws_summary._current_row, column=5).column_letter].width = 20


  totalcost_column_index = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']

  spot_pyxl = dataframe_to_rows(spotusage)
  gcpcost_column_letter = 'Z'
  for r_idx, row in enumerate(spot_pyxl, 1):
      for c_idx, value in enumerate(row, 1):
        if value==columnnames['cost']:
          totalcost_column_letter = totalcost_column_index[c_idx-1]
        if value=='gcp_cost':
          gcpcost_column_letter = totalcost_column_index[c_idx-1]
        ws_spot.cell(row=r_idx, column=c_idx, value=value)
  ws_summary.cell(row=ws_summary._current_row, column=4).value = '=SUM(SpotUsage!{}:{})'.format(gcpcost_column_letter, gcpcost_column_letter)
  ws_summary.cell(row=ws_summary._current_row, column=6).value = '=SUM(SpotUsage!{}:{})'.format(totalcost_column_letter, totalcost_column_letter)

  ws_summary.append(['', 'Compute', 'GCE - Regular VMs(Box)', '${}'.format(cost_matrix['gcp']['box_compute']), 'EC2 - Regular Usage(Box)', '${}'.format(cost_matrix['aws']['box_compute'])])
  ws_box = wb.create_sheet(title='BoxUsage')
  box_pyxl = dataframe_to_rows(boxusage)
  gcpcost_column_letter = 'Z'
  for r_idx, row in enumerate(box_pyxl, 1):
      for c_idx, value in enumerate(row, 1):
        if value==columnnames['cost']:
          totalcost_column_letter = totalcost_column_index[c_idx-1]
        if value=='gcp_cost':
          gcpcost_column_letter = totalcost_column_index[c_idx-1]
        ws_box.cell(row=r_idx, column=c_idx, value=value)
  ws_summary.cell(row=ws_summary._current_row, column=4).value = '=SUM(BoxUsage!{}:{})'.format(gcpcost_column_letter, gcpcost_column_letter)
  ws_summary.cell(row=ws_summary._current_row, column=6).value = '=SUM(BoxUsage!{}:{})'.format(totalcost_column_letter, totalcost_column_letter)

  ws_summary.append(['', 'Compute', 'GCE - Regular VMs(Heavy)', '${}'.format(cost_matrix['gcp']['heavy_compute']), 'EC2 - Regular Usage(Heavy)', '${}'.format(cost_matrix['aws']['heavy_compute'])])
  ws_heavy = wb.create_sheet(title='HeavyUsage')
  heavy_pyxl = dataframe_to_rows(heavyusage)
  gcpcost_column_letter = 'Z'
  for r_idx, row in enumerate(heavy_pyxl, 1):
      for c_idx, value in enumerate(row, 1):
        if value==columnnames['cost']:
          totalcost_column_letter = totalcost_column_index[c_idx-1]
        if value=='gcp_cost':
          gcpcost_column_letter = totalcost_column_index[c_idx-1]
        ws_heavy.cell(row=r_idx, column=c_idx, value=value)
  ws_summary.cell(row=ws_summary._current_row, column=4).value = '=SUM(HeavyUsage!{}:{})'.format(gcpcost_column_letter, gcpcost_column_letter)
  ws_summary.cell(row=ws_summary._current_row, column=6).value = '=SUM(HeavyUsage!{}:{})'.format(totalcost_column_letter, totalcost_column_letter)

  ws_summary.append(['', '', '', '${}'.format(cost_matrix['gcp']['box_compute']+cost_matrix['gcp']['heavy_compute']+cost_matrix['gcp']['spot_compute']), '', '${}'.format(cost_matrix['aws']['box_compute']+cost_matrix['aws']['heavy_compute']+cost_matrix['aws']['spot_compute'])])
  ws_summary.cell(row=ws_summary._current_row, column=4).value = '=SUM(D4:D6)'
  ws_summary.cell(row=ws_summary._current_row, column=6).value = '=SUM(F4:F6)'

  ws_summary.append(['', 'Storage', 'Persistent Disk', '${}'.format(cost_matrix['gcp']['persistentdisk']), 'Elastic Block Storage', '${}'.format(cost_matrix['aws']['persistentdisk'])])
  ws_pd = wb.create_sheet(title='Persistent Disk')
  pd_pyxl = dataframe_to_rows(persistentdisk)
  for r_idx, row in enumerate(pd_pyxl, 1):
      for c_idx, value in enumerate(row, 1):
        if value==columnnames['cost']:
          totalcost_column_letter = totalcost_column_index[c_idx-1]
        ws_pd.cell(row=r_idx, column=c_idx, value=value)
  ws_summary.cell(row=ws_summary._current_row, column=4).value = "=SUM('Persistent Disk'!I:I)"
  ws_summary.cell(row=ws_summary._current_row, column=6).value = "=SUM('Persistent Disk'!{}:{})".format(totalcost_column_letter, totalcost_column_letter)

  ws_summary.append(['', 'Storage', 'Cloud Storage', '${}'.format(cost_matrix['gcp']['cloudstorage']), 'Simple Storage Service(S3)', '${}'.format(cost_matrix['aws']['cloudstorage'])])
  ws_gcs = wb.create_sheet(title='Cloud Storage')
  gcs_pyxl = dataframe_to_rows(cloudstorage)
  for r_idx, row in enumerate(gcs_pyxl, 1):
      for c_idx, value in enumerate(row, 1):
        if value==columnnames['cost']:
          totalcost_column_letter = totalcost_column_index[c_idx-1]
        ws_gcs.cell(row=r_idx, column=c_idx, value=value)
  ws_summary.cell(row=ws_summary._current_row, column=4).value = "=SUM('Cloud Storage'!L:L)"
  ws_summary.cell(row=ws_summary._current_row, column=6).value = "=SUM('Cloud Storage'!{}:{})".format(totalcost_column_letter, totalcost_column_letter)
  # ws_summary.append(['', 'Storage', 'Filestore', '', 'Elastic File Storage', ''])
  ws_summary.append(['', '', '', '', '', ''])
  ws_summary.cell(row=ws_summary._current_row, column=4).value = '=SUM(D8:D9)'
  ws_summary.cell(row=ws_summary._current_row, column=6).value = '=SUM(F8:F9)'

  ws_summary.append(['', 'Networking', 'Cloud Load Balancer', '${}'.format(cost_matrix['gcp']['loadbalancer']), 'Elastic Load Balancer', '${}'.format(cost_matrix['aws']['loadbalancer'])])
  ws_lb = wb.create_sheet(title='Load Balancer')
  lb_pyxl = dataframe_to_rows(loadbalancer)
  for r_idx, row in enumerate(lb_pyxl, 1):
      for c_idx, value in enumerate(row, 1):
        if value==columnnames['cost']:
          totalcost_column_letter = totalcost_column_index[c_idx-1]
        ws_lb.cell(row=r_idx, column=c_idx, value=value)
  ws_summary.cell(row=ws_summary._current_row, column=4).value = "=SUM('Load Balancer'!L:L)"
  ws_summary.cell(row=ws_summary._current_row, column=6).value = "=SUM('Load Balancer'!{}:{})".format(totalcost_column_letter, totalcost_column_letter)

  ws_summary.append(['', 'Networking', 'Cloud NAT', '${}'.format(cost_matrix['gcp']['cloudnat']), 'NAT Gateway', '${}'.format(cost_matrix['aws']['cloudnat'])])

  ws_summary.append(['', 'Networking', 'Network Egress', '${}'.format(cost_matrix['gcp']['egress']), 'Data Transfer', '${}'.format(cost_matrix['aws']['egress'])])
  ws_negress = wb.create_sheet(title='Network Egress')
  negress_pyxl = dataframe_to_rows(egress_df)
  for r_idx, row in enumerate(negress_pyxl, 1):
      for c_idx, value in enumerate(row, 1):
        if value==columnnames['cost']:
          totalcost_column_letter = totalcost_column_index[c_idx-1]
        ws_negress.cell(row=r_idx, column=c_idx, value=value)
  ws_summary.cell(row=ws_summary._current_row, column=4).value = "=SUM('Network Egress'!G:G)"
  ws_summary.cell(row=ws_summary._current_row, column=6).value = "=SUM('Network Egress'!{}:{})".format(totalcost_column_letter, totalcost_column_letter)

  ws_summary.append(['', 'Networking', 'Idle Addresseses', '${}'.format(cost_matrix['gcp']['idleaddress']), 'Idle Addresses', '${}'.format(cost_matrix['aws']['idleaddress'])])
  ws_idadress = wb.create_sheet(title='Idle Addresses')
  Idaddress_pyxl = dataframe_to_rows(idleaddress)
  for r_idx, row in enumerate(Idaddress_pyxl, 1):
      for c_idx, value in enumerate(row, 1):
        if value==columnnames['cost']:
          totalcost_column_letter = totalcost_column_index[c_idx-1]
        ws_idadress.cell(row=r_idx, column=c_idx, value=value)
  ws_summary.cell(row=ws_summary._current_row, column=4).value = "=SUM('Idle Addresses'!H:H)"
  ws_summary.cell(row=ws_summary._current_row, column=6).value = "=SUM('Idle Addresses'!{}:{})".format(totalcost_column_letter, totalcost_column_letter)

  ws_summary.append(['', '', '', '', '', ''])
  ws_summary.cell(row=ws_summary._current_row, column=4).value = '=SUM(D11:D14)'
  ws_summary.cell(row=ws_summary._current_row, column=6).value = '=SUM(F11:F14)'

  ws_cloudsql = wb.create_sheet(title='Cloud SQL')
  cloudsql_pyxl = dataframe_to_rows(cloudsql)
  for r_idx, row in enumerate(cloudsql_pyxl, 1):
      for c_idx, value in enumerate(row, 1):
        if value==columnnames['cost']:
          totalcost_column_letter = totalcost_column_index[c_idx-1]
        ws_cloudsql.cell(row=r_idx, column=c_idx, value=value)
  # ws_summary.append(['', 'DB Services', 'Cloud SQL','', 'Amazon RDS', ''])
  ws_summary.append(['', 'DB Services', 'Cloud SQL', '${}'.format(cost_matrix['gcp']['cloudsql']), 'Amazon RDS', '${}'.format(cost_matrix['aws']['cloudsql'])])
  ws_summary.cell(row=ws_summary._current_row, column=4).value = "=SUM('Cloud SQL'!I:I)"
  # ws_summary.append(['', 'DB Services', 'Search on GCP', '', 'ElasticSearch', ''])
  # ws_summary.append(['', 'DB Services', 'Cache on GCP', '', 'Elasticache', ''])
  # ws_summary.append(['', 'DB Services', 'BigQuery', '', 'Redshit', ''])
  ws_summary.append(['', '', '', '', '', ''])
  ws_summary.cell(row=ws_summary._current_row, column=4).value = '=SUM(D16:D16)'
  ws_summary.cell(row=ws_summary._current_row, column=6).value = '=SUM(F16:F16)'

  ws_summary.append(['', 'Support', 'GCP Support', 0, 'AWS Support Business', sum(support[columnnames['cost']])])

  ws_summary.append(['', 'Misc', 'Unclassified', '', 'Misc', ''])
  ws_summary.cell(row=ws_summary._current_row, column=4).value = "=SUM('Uncomputed'!{}:{})".format(totalcost_column_letter, totalcost_column_letter)
  ws_summary.cell(row=ws_summary._current_row, column=6).value = "=SUM('Uncomputed'!{}:{})".format(totalcost_column_letter, totalcost_column_letter)

  ws_summary.append(['', '', '', '', '', ''])
  ws_summary.cell(row=ws_summary._current_row, column=4).value = '=SUM(D18:D19)'
  ws_summary.cell(row=ws_summary._current_row, column=6).value = '=SUM(F18:F19)'

  ws_summary.append(['', 'GCP Monthly', '', '', 'AWS Monthly', '', ''])
  ws_summary.cell(row=ws_summary._current_row, column=4).value = '=D7+D10+D15+D17+D20'
  ws_summary.cell(row=ws_summary._current_row, column=6).value = '=F7+F10+F15+F17+F20'

  #Throw everything else into the 'Other' tab
  border = Border(left=Side(border_style='thin', color='000000'),
                  right=Side(border_style='thin', color='000000'),
                  top=Side(border_style='thin', color='000000'),
                  bottom=Side(border_style='thin', color='000000'))

  for i in range(4, 22):
    ws_summary.cell(row=i, column=7).value = "=1-D{}/F{}".format(i, i)
    ws_summary.cell(row=i, column=7).number_format = "0.00%"
    ws_summary.cell(row=i, column=4).number_format = "$0.00"
    ws_summary.cell(row=i, column=6).number_format = "$0.00"
    for col in range(2, 8):
      ws_summary.cell(row=i, column=col).border = border

  ws_others = wb.create_sheet(title='Uncomputed')
  others_pyxl = dataframe_to_rows(grouped)
  for r_idx, row in enumerate(others_pyxl, 1):
      for c_idx, value in enumerate(row, 1):
        if value==columnnames['cost']:
          totalcost_column_letter = totalcost_column_index[c_idx-1]
        ws_others.cell(row=r_idx, column=c_idx, value=value)

  # def set_border(ws, cell_range):
  #     border = Border(left=Side(border_style='thin', color='000000'),
  #                 right=Side(border_style='thin', color='000000'),
  #                 top=Side(border_style='thin', color='000000'),
  #                 bottom=Side(border_style='thin', color='000000'))

  #     rows = ws.iter_rows(cell_range)
  #     for row in rows:
  #         for cell in row:
  #             cell.border = border

  # set_border(ws_summary, 'B2:G21')


  wb.save(filename=report_filename)

def main():
  tcocomputer()
  reporting()
######################################################################################################################################################

print('Remaining ', sum(grouped[columnnames['cost']]))